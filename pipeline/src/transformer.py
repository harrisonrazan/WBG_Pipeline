# pipeline/src/transformer.py
"""Functions for transforming and processing World Bank data into structured formats"""

import pandas as pd
from typing import Dict, Any, List
from datetime import datetime
import re
import logging
from config import TABLES
import csv
import io
import openpyxl
import concurrent.futures

logger = logging.getLogger(__name__)

def process_api_call_json(data: List[Dict[str, Any]], api_endpoint: str) -> pd.DataFrame:
    try:
        logger.info(f"Processing {api_endpoint} data...")
        df = pd.DataFrame(data)
        df.columns = [standardize_column_name(col) for col in df.columns]
        if 'processed_at' not in df.columns and 'as_of_date' not in df.columns:
            df['as_of_date'] = datetime.now()
        return df
    except Exception as e:
        logger.error(f"Error processing {api_endpoint}:{str(e)}")


def standardize_column_name(column: str) -> str:
    """
    Standardizes column names to lowercase with underscores.
    
    Args:
        column: Original column name from Excel
        
    Returns:
        Standardized column name in lowercase with underscores
    """
    # Convert to lowercase
    name = str(column).lower()
    
    # Replace special characters and spaces with underscore
    import re
    # Replace any non-alphanumeric character (except underscores) with underscore
    name = re.sub(r'[^a-z0-9_]', '_', name)
    
    # Replace multiple consecutive underscores with a single underscore
    name = re.sub(r'_+', '_', name)
    
    # Remove leading/trailing underscores
    name = name.strip('_')
    
    return name

def process_excel_worker(sheet_info):
    """Worker function to process a single Excel sheet - defined outside for pickling"""
    import logging
    import pandas as pd
    from datetime import datetime
    import openpyxl
    from transformer import standardize_column_name
    
    logger = logging.getLogger(__name__)
    file_path, sheet_name, skiprows = sheet_info
    
    try:
        logger.info(f"Processing sheet '{sheet_name}'...")
        
        # Read the Excel sheet with pandas, skipping specified rows
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            skiprows=skiprows
        )
        
        # Standardize column names
        df.columns = [standardize_column_name(col) for col in df.columns]
        
        # Get the actual sheet object from openpyxl
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        
        # Check if project_id or id column exists
        project_id_col = None
        for col in df.columns:
            if col in ['project_id', 'id'] or 'project' in col and 'id' in col:
                project_id_col = col
                break
        
        # If we found a project ID column, extract hyperlinks
        if project_id_col is not None:
            # Find the column index in the Excel sheet
            col_idx = None
            for i, col_name in enumerate(df.columns):
                if col_name == project_id_col:
                    col_idx = i
                    break
            
            if col_idx is not None:
                # Create a new column for URLs
                url_col_name = f"{project_id_col}_url"
                df[url_col_name] = None
                
                # Extract hyperlinks from each cell in the project_id column
                # Add offset for skipped rows and header
                row_offset = 1 + len(skiprows)
                
                for excel_row_idx, df_row_idx in enumerate(range(len(df)), start=row_offset):
                    cell = sheet.cell(row=excel_row_idx, column=col_idx + 1)  # +1 because Excel is 1-indexed
                    if cell.hyperlink:
                        df.at[df_row_idx-1, url_col_name] = cell.hyperlink.target
        
        # Convert date columns if present (using standardized column names)
        date_columns = [col for col in df.columns
                      if any(date_term in col
                            for date_term in ['date', 'as_of'])]
        for col in date_columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        # Convert numeric columns (using standardized column names)
        numeric_columns = [col for col in df.columns
                         if any(amount_term in col
                               for amount_term in ['amount', 'cost', 'commitment'])]
        for col in numeric_columns:
            # Convert string with commas to numeric
            if df[col].dtype == object:  # Check if it's a string column
                df[col] = df[col].astype(str).str.replace(',', '')
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Add processed timestamp
        if 'processed_at' not in df.columns and 'as_of_date' not in df.columns:
            df['as_of_date'] = datetime.now()
        
        # Store the DataFrame using the sheet name as key
        # Convert sheet name to lowercase and replace spaces with underscores
        sheet_key = standardize_column_name(sheet_name)
        
        logger.info(f"Processed sheet '{sheet_name}' with {len(df)} rows")
        return (sheet_key, df)
        
    except Exception as e:
        logger.error(f"Error processing Excel sheet '{sheet_name}': {str(e)}")
        return (standardize_column_name(sheet_name), None)
    
def process_projects_excel_concurrently(file_path):
    """
    Processes the World Bank projects Excel file into multiple DataFrames concurrently.
    Uses ProcessPoolExecutor for CPU-intensive processing.
    """
    try:
        # Define sheets to process
        sheet_names = ['World Bank Projects', 'Themes', 'Sectors', 'GEO Locations', 'Financers']
        
        # Prepare tasks for concurrent execution
        sheet_tasks = []
        for sheet_name in sheet_names:
            # For World Bank Projects sheet, skip rows 0 and 2 (indices)
            # For all other sheets, skip only row 0
            skiprows = [0, 2] if sheet_name == 'World Bank Projects' else [0]
            sheet_tasks.append((file_path, sheet_name, skiprows))
        
        dataframes = {}
        
        # Use ProcessPoolExecutor for CPU-intensive work
        with concurrent.futures.ProcessPoolExecutor() as executor:
            # Submit all tasks
            future_to_sheet = {executor.submit(process_excel_worker, task): task[1] for task in sheet_tasks}
            
            # Process results as they complete
            for future in concurrent.futures.as_completed(future_to_sheet):
                sheet_name = future_to_sheet[future]
                try:
                    sheet_key, df = future.result()
                    if df is not None:
                        dataframes[sheet_key] = df
                        logging.debug(f"Standardized columns for '{sheet_name}': {list(df.columns)}")
                except Exception as e:
                    logging.error(f"Exception processing sheet '{sheet_name}': {str(e)}")
                    
        return dataframes
    except Exception as e:
        logging.error(f"Error in concurrent Excel processing: {str(e)}")
        raise
    
def process_projects_excel(file_path: str) -> Dict[str, pd.DataFrame]:
    """
    Processes the World Bank projects Excel file into multiple DataFrames.
    Preserves hyperlinks in the project_id column.
    """
    try:
        # Read all sheets from Excel file
        sheet_names = ['World Bank Projects', 'Themes', 'Sectors',
                      'GEO Locations', 'Financers']
       
        dataframes = {}
        
        # Load the workbook using openpyxl to access hyperlinks
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
       
        for sheet_name in sheet_names:
            # For World Bank Projects sheet, skip rows 0 and 2 (indices)
            # For all other sheets, skip only row 0
            skiprows = [0, 2] if sheet_name == 'World Bank Projects' else [0]
           
            # Read the Excel sheet with pandas, skipping specified rows
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                skiprows=skiprows
            )
           
            # Standardize column names
            df.columns = [standardize_column_name(col) for col in df.columns]
            
            # Get the actual sheet object from openpyxl
            sheet = workbook[sheet_name]
            
            # Check if project_id or id column exists
            project_id_col = None
            for col in df.columns:
                if col in ['project_id', 'id'] or 'project' in col and 'id' in col:
                    project_id_col = col
                    break
            
            # If we found a project ID column, extract hyperlinks
            if project_id_col is not None:
                # Find the column index in the Excel sheet
                col_idx = None
                for i, col_name in enumerate(df.columns):
                    if col_name == project_id_col:
                        col_idx = i
                        break
                
                if col_idx is not None:
                    # Create a new column for URLs
                    url_col_name = f"{project_id_col}_url"
                    df[url_col_name] = None
                    
                    # Extract hyperlinks from each cell in the project_id column
                    # Add offset for skipped rows and header
                    row_offset = 1 + len(skiprows)
                    
                    for excel_row_idx, df_row_idx in enumerate(range(len(df)), start=row_offset):
                        cell = sheet.cell(row=excel_row_idx, column=col_idx + 1)  # +1 because Excel is 1-indexed
                        if cell.hyperlink:
                            df.at[df_row_idx-1, url_col_name] = cell.hyperlink.target
           
            # Convert date columns if present (using standardized column names)
            date_columns = [col for col in df.columns
                          if any(date_term in col
                                for date_term in ['date', 'as_of'])]
            for col in date_columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
           
            # Convert numeric columns (using standardized column names)
            numeric_columns = [col for col in df.columns
                             if any(amount_term in col
                                   for amount_term in ['amount', 'cost', 'commitment'])]
            for col in numeric_columns:
                # Convert string with commas to numeric
                if df[col].dtype == object:  # Check if it's a string column
                    df[col] = df[col].astype(str).str.replace(',', '')
                df[col] = pd.to_numeric(df[col], errors='coerce')
           
            # Store the DataFrame using the sheet name as key
            # Convert sheet name to lowercase and replace spaces with underscores
            sheet_key = standardize_column_name(sheet_name)
            if 'processed_at' not in df.columns and 'as_of_date' not in df.columns:
                df['as_of_date'] = datetime.now()
            dataframes[sheet_key] = df
           
            logging.info(f"Processed sheet '{sheet_name}' with {len(df)} rows")
            logging.debug(f"Standardized columns for '{sheet_name}': {list(df.columns)}")
       
        return dataframes
    except Exception as e:
        logging.error(f"Error processing Excel file: {str(e)}")
        raise

def process_gef_projects_csv(file_path: str) -> pd.DataFrame:
    try:
        # Read the CSV file, handling potential issues
        df = pd.read_csv(
            file_path
            # engine='python',           # Use more flexible parser
            # encoding='utf-8-sig',      # Handle BOM if present
            # quoting=csv.QUOTE_MINIMAL, # Only quote when needed
            # escapechar='\\',           # Define escape character
            # skipinitialspace=True      # Skip spaces after delimiter
        )
        
        logger.info(f"Successfully parsed GEF projects CSV file with {len(df)} rows")
        
        # Standardize column names
        df.columns = [standardize_column_name(col) for col in df.columns]
        
        if 'processed_at' not in df.columns and 'as_of_date' not in df.columns:
            df['as_of_date'] = datetime.now()
        
        return df
    except Exception as e:
        logger.error(f"Error processing GEF projects CSV file: {str(e)}")
        raise